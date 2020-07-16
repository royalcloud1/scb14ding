# -*- coding: utf-8 -*-
# @Time ： 2020/7/15 20:28
# @Auth ： jianwen
# @File ：lianxi5.py
# @QQ ：673223601
#'test_case_api.xlsx','login'

import requests
import openpyxl

def fun_request(url,data):
    headers_request = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
    response = requests.post(url = url,json=data,headers=headers_request)
    res = response.json()
    return res

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    list1 = []
    for i in range(2,max_row+1):
        dict1=dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expected = sheet.cell(row=i,column=7).value,
        )
        list1.append(dict1)
    return list1

def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

def exceted_fun(filename,sheetname):
    cases = read_data(filename,sheetname)
    for i in cases:
        case_id = i.get('case_id')
        url = i.get('url')
        data = eval(i.get('data'))
        expected = eval(i.get('expected'))
        expected_msg = expected.get('msg')
        expected_code = expected.get('code')

        real_result = fun_request(url=url,data=data)
        real_msg = real_result.get('msg')
        real_code = real_result.get('code')

        print('预期结果的code：{}'.format(expected_code))
        print('实际结果的code：{}'.format(real_code))
        print('预期结果的msg：{}'.format(expected_msg))
        print('实际结果的msg：{}'.format(real_msg))
        if expected_msg==real_msg and expected_code==real_code:
            print('第{}条用例执行通过'.format(case_id))
            final_re = 'XXX'
        else:
            print('第{}条用例执行不通过'.format(case_id))
            final_re = 'YYY'
        write_result(filename,sheetname,case_id+1,8,final_re)
        print('*'*20)
exceted_fun('test_case_api.xlsx','login')


